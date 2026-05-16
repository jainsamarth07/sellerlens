"""Generate a Flipkart-shaped settlement workbook for sellers to download as a
reference for the expected upload format.

Mirrors the real Flipkart Marketplace settlement export structure:
  - "Summary of report"  : structured 'Line Item / Break-up L1 / Break-up L2 /
                           Amount Settled (Rs.) / …' table.
  - "Orders"             : multi-row header (group row + column row) with
                           '(Rs.)' suffixes — production has 76 columns; we
                           expose the most-used subset.
  - "Ads"                : group-header row + 'Type / Wallet Topup / Wallet
                           Refund / GST on Ads Fees' columns.
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta

from openpyxl import Workbook


def build_sample_flipkart_workbook() -> bytes:
    """Return a Flipkart-shaped .xlsx file the parser can fully ingest."""
    wb = Workbook()

    _write_summary_sheet(wb)
    _write_orders_sheet(wb)
    _write_ads_sheet(wb)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary of report")

    ws.append([None, "Settlement Report Summary"])
    ws.append([None, "NOTE: Sample settlement report — replace with your real Flipkart export."])
    ws.append([None, None])
    ws.append([None, "Payment Duration ", "01-Apr-2026 - 30-Apr-2026"])
    ws.append([None, "# Sale Orders (Not returned)", 124])
    ws.append([None, "# Returns", 18])
    ws.append([None, None])
    ws.append([
        None, "Line Item", "Break-up L1", "Break-up L2",
        "Amount Settled (Rs.)", "Break-up L1 (Rs.)", "Break-up L2 (Rs.)",
        "Remarks (Information about each line)",
    ])
    ws.append([None, "Orders", None, None, 197295.00])
    ws.append([None, None, "Sales Amount", None, None, 285000.00, None,
               "Sum of sales amount paid by customers in the selected period"])
    ws.append([None, None, "Returns Reversal", None, None, -41200.00, None,
               "Sum of sales amount reversed due to returns"])
    ws.append([None, None, "Marketplace Fees", None, None, -38500.00, None,
               "Sum of Marketplace fees for sales & returned orders"])
    ws.append([None, None, "Taxes", None, -11205.00, None, None,
               "Tax breakup below"])
    ws.append([None, None, None, "TCS", None, None, -2850.00,
               "TCS amount eligible for input tax credit"])
    ws.append([None, None, None, "TDS", None, None, -1425.00,
               "TDS deducted, eligible for income-tax reimbursement"])
    ws.append([None, None, None, "GST on marketplace fees", None, None, -6930.00,
               "Service GST on Marketplace fees, eligible for input GST credit"])
    ws.append([None, "Protection Fund (SPF)", None, None, 0.00])
    ws.append([None, "Services Fees", None, None, -14160.00])
    ws.append([None, None, "Ads Fees", None, None, -12000.00, None,
               "Sum of ad-wallet transactions"])
    ws.append([None, None, "Taxes", None, -2160.00, None, None,
               "Service tax breakup"])
    ws.append([None, None, None, "GST on Ads Fees", None, None, -2160.00,
               "Service GST on Ads campaign fees"])
    ws.append([None, "Tax Settlement", None, None, 0.00])
    ws.append([None, "Net Bank Settlement (A)", None, None, 182095.00, None, None,
               "Amount received from Flipkart as bank settlement"])
    ws.append([None, "Input GST + TCS Credits (B)", None, None, 9780.00, None, None,
               "Eligible for Input Tax Credit during GSTR filing"])
    ws.append([None, "Income Tax Credits (C)", None, None, 1425.00, None, None,
               "Eligible for TDS reimbursement during income tax filing"])
    ws.append([None, "Total Realizable Amount", None, None, 193300.00, None, None,
               "Net Bank Settlement (A) + Input GST credits (B) + Income Tax Credits (C)"])


# ---------------------------------------------------------------------------
# Orders sheet
# ---------------------------------------------------------------------------

_ORDERS_GROUP_ROW = (
    ["Payment Details"] * 6
    + ["Order Details"] * 5
    + ["Financials"] * 5
    + ["Fees"] * 7
    + ["Tax"] * 3
    + ["Product"] * 4
)

_ORDERS_COLUMN_ROW = [
    "NEFT ID", "Neft Type", " Payment Date",
    "Bank Settlement Value (Rs.) \n= SUM(J:R)",
    "Input GST + TCS Credits (Rs.)\n[GST+TCS]",
    "Income Tax Credits (Rs.)\n[TDS]",
    "Order ID", "Order item ID", "Order Date", "Dispatch Date", "Fulfilment Type",
    "Sale Amount (Rs.)", "Total Offer Amount (Rs.)", "Marketplace Fee (Rs.)\n= SUM (V:AI)",
    "Taxes (Rs.)", "Refund (Rs.)",
    "Commission (Rs.)", "Fixed Fee  (Rs.)", "Collection Fee (Rs.)",
    "Pick And Pack Fee (Rs.)", "Shipping Fee (Rs.)", "Reverse Shipping Fee (Rs.)",
    "Protection Fund (Rs.)",
    "TCS (Rs.)", "TDS (Rs.)", "GST on MP Fees (Rs.)",
    "Seller SKU", "Quantity", "Product Sub Category", "Return Type",
]


def _write_orders_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Orders")
    ws.append(_ORDERS_GROUP_ROW)
    ws.append(_ORDERS_COLUMN_ROW)

    base_date = datetime(2026, 4, 1)
    skus = [
        ("SKU-001-RED-M", "Apparel", 1500),
        ("SKU-001-BLU-L", "Apparel", 1750),
        ("SKU-002-BLK-S", "Home", 800),
        ("SKU-003-WHT-XL", "Footwear", 2200),
    ]
    for i in range(20):
        sku, category, base_price = skus[i % len(skus)]
        is_return = i in {3, 8, 14, 17}
        sale = base_price + (i % 3) * 100
        mp_fee = -round(sale * 0.13, 2)
        taxes = -round(sale * 0.05, 2)
        refund = -sale if is_return else 0
        bank_settlement = 0 if is_return else round(sale + mp_fee + taxes, 2)
        commission = -round(sale * 0.08, 2)
        rev_shipping = -60 if is_return else 0
        tcs = -round(sale * 0.01, 2)
        tds = -round(sale * 0.005, 2)
        gst_mp = -round(abs(mp_fee) * 0.18, 2)
        date_str = (base_date + timedelta(days=i)).strftime("%Y-%m-%d")
        ws.append([
            f"NFT-SAMPLE-{i:04d}", "PREPAID", date_str,
            bank_settlement, abs(tcs) + abs(gst_mp), abs(tds),
            f"OD{1000 + i}", f"OI{2000 + i}",
            (base_date + timedelta(days=i - 2)).strftime("%Y-%m-%d"),
            (base_date + timedelta(days=i - 1)).strftime("%Y-%m-%d"),
            "Flipkart Smart",
            sale, 0, mp_fee, taxes, refund,
            commission, -25, -12,
            -10, -45, rev_shipping, 0,
            tcs, tds, gst_mp,
            sku, 1, category,
            "Customer Return" if is_return else None,
        ])


# ---------------------------------------------------------------------------
# Ads sheet
# ---------------------------------------------------------------------------

def _write_ads_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Ads")
    ws.append([
        "Payment Details", None, None, None,
        "Transaction Summary", None, None, None, None, None, None,
    ])
    ws.append([
        "NEFT ID", "Payment Date", "Settlement Value (Rs.) = SUM(G:K)", None,
        "Type", "Campaign / Transaction ID",
        "Wallet Redeem (Rs.)", "Wallet Redeem Reversal (Rs.)",
        "Wallet Topup (Rs.)", "Wallet Refund (Rs.)", "GST on Ads Fees (Rs.)",
    ])
    base_date = datetime(2026, 4, 5)
    rows = [
        ("topup",  "TXN-260405-T1", 0,      0, 8000, 0, -1440),
        ("redeem", "PCID-260410-A1", -3200, 0, 0,    0, -576),
        ("redeem", "PCID-260415-A2", -2800, 0, 0,    0, -504),
        ("topup",  "TXN-260420-T2", 0,      0, 4000, 0, -720),
        ("redeem", "PCID-260425-A3", -1900, 0, 0,    0, -342),
    ]
    for i, (txn_type, cid, redeem, redeem_rev, topup, refund, gst) in enumerate(rows):
        date_str = (base_date + timedelta(days=i * 5)).strftime("%Y-%m-%d")
        settlement = topup - refund
        ws.append([
            f"NFT-ADS-{i:04d}", date_str, settlement, None,
            txn_type, cid,
            redeem, redeem_rev, topup, refund, gst,
        ])
