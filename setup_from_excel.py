"""
Read Project_Files.xlsx and recreate the entire SellerLens project structure.

This script reads all 3 sheets (Frontend, Backend, Other Files) from the
Excel workbook and creates every file at its correct relative path,
preserving directory structure.

Usage:
    python setup_from_excel.py                          # creates in current directory
    python setup_from_excel.py --target C:\my\project   # creates in specified directory
    python setup_from_excel.py --excel path/to/file.xlsx # use a different Excel file
"""

import argparse
import base64
import sys
from pathlib import Path

from openpyxl import load_workbook


def setup_project(excel_path: Path, target_dir: Path):
    """Read the Excel workbook and recreate all files."""
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"Reading : {excel_path}")
    print(f"Target  : {target_dir}\n")

    wb = load_workbook(str(excel_path), read_only=True, data_only=True)

    total_created = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_created = 0
        sheet_skipped = 0

        print(f"--- Sheet: {sheet_name} ---")

        first_row = True
        for row in ws.iter_rows(min_row=1, values_only=True):
            # Skip header row
            if first_row:
                first_row = False
                continue

            # Expect columns: #, File Path, File Content, Encoded Content
            if len(row) < 3:
                continue

            _, file_path_str, plain_content, *rest = row
            encoded_content = rest[0] if rest else None

            if not file_path_str:
                continue

            # Normalise path
            rel_path = Path(str(file_path_str).strip())
            full_path = target_dir / rel_path

            # Prefer base64-encoded column (exact fidelity), fall back to plain
            if encoded_content:
                try:
                    content = base64.b64decode(str(encoded_content)).decode("utf-8")
                except Exception:
                    content = plain_content if plain_content is not None else ""
            else:
                content = plain_content if plain_content is not None else ""

            try:
                full_path.parent.mkdir(parents=True, exist_ok=True)
                full_path.write_text(str(content), encoding="utf-8")
                sheet_created += 1
            except Exception as e:
                print(f"  SKIP  {rel_path}  ({e})")
                sheet_skipped += 1

        print(f"  Created: {sheet_created}  |  Skipped: {sheet_skipped}")
        total_created += sheet_created
        total_skipped += sheet_skipped

    wb.close()

    print(f"\n{'='*50}")
    print(f"Done!  Files created : {total_created}")
    if total_skipped:
        print(f"       Files skipped : {total_skipped}")
    print(f"       Target folder : {target_dir}")
    print()
    print("Next steps:")
    print("  1. cp .env.example .env  (fill in Azure keys)")
    print("  2. pip install -r requirements.txt")
    print("  3. cd frontend && npm install")
    print("  4. See README.md for full setup instructions")


def main():
    parser = argparse.ArgumentParser(
        description="Recreate SellerLens project files from Project_Files.xlsx"
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="Project_Files.xlsx",
        help="Path to the Excel workbook (default: Project_Files.xlsx in current dir)",
    )
    parser.add_argument(
        "--target",
        type=str,
        default=".",
        help="Target directory to create files in (default: current directory)",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    target_dir = Path(args.target).resolve()

    setup_project(excel_path, target_dir)


if __name__ == "__main__":
    main()
