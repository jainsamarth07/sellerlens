"""Unit tests for the Flipkart and Amazon settlement parsers."""

from __future__ import annotations

import io

import pandas as pd
import pytest

from backend.processors._helpers import to_float, to_int
from backend.processors.amazon_parser import parse_amazon_settlement
from backend.processors.flipkart_parser import (
    aggregate_skus,
    parse_ads_sheet,
    parse_flipkart_settlement,
    parse_orders_sheet,
    parse_summary_sheet,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_flipkart_workbook() -> bytes:
    """Build a minimal in-memory Flipkart settlement workbook using openpyxl."""
    from openpyxl import Workbook

    wb = Workbook()
    # Default sheet → Summary
    summary_ws = wb.active
    summary_ws.title = "Summary of report"
    summary_rows = [
        ["Payment Duration", "01-Apr to 15-Apr"],
        ["Total Sale Orders", 120],
        ["Total Returns", 8],
        ["Sales Amount", "₹1,50,000.00"],
        ["Returns Reversal", "-12000"],
        ["Marketplace Fees", "-15000"],
        ["TCS Amount", "-1500"],
        ["TDS Amount", "-1000"],
        ["GST on MP Fees", "-2700"],
        ["Ads Fees", "5000"],
        ["Net Bank Settlement", "112800"],
        ["Input GST TCS Credits", "2700"],
        ["Income Tax Credits", "1000"],
        ["Total Realizable Amount", "116500"],
    ]
    for row in summary_rows:
        summary_ws.append(row)

    # Orders sheet — row 0 = group, row 1 = column names, then data
    orders_ws = wb.create_sheet("Orders")
    orders_cols = [
        "Order ID", "Order Item ID", "Payment Date",
        "Sale Amount", "Marketplace Fee", "Taxes",
        "Refund Amount", "Bank Settlement Value", "Seller SKU ID",
        "Quantity", "Product Sub Category",
        "Commission", "Fixed Fee", "Collection Fee", "Reverse Shipping Fee",
        "TCS", "TDS", "GST on MP Fees",
    ]
    group_row = ["Order Info"] * 3 + ["Financials"] * 6 + ["Product"] * 2 + ["Fees"] * 7
    orders_ws.append(group_row)
    orders_ws.append(orders_cols)
    data_rows = [
        ["OD111", "OI111", "2026-04-05", 1500, -150, 270, 0, 1350, "SKU-A", 1, "Apparel", -100, -10, -5, 0, -15, -10, -27],
        ["OD112", "OI112", "2026-04-06", 800, -80, 144, 0, 720, "SKU-B", 2, "Home", -60, -5, -3, 0, -8, -5, -14],
        ["OD113", "OI113", "2026-04-07", 1500, -150, 270, -1500, -150, "SKU-A", 1, "Apparel", 0, 0, 0, -50, 0, 0, 0],
    ]
    for row in data_rows:
        orders_ws.append(row)

    # Ads sheet
    ads_ws = wb.create_sheet("Ads")
    ads_ws.append(["Campaign ID", "Transaction Type", "Amount", "GST on Ads"])
    ads_ws.append(["CMP1", "Spend", 3000, 540])
    ads_ws.append(["CMP2", "Spend", 2000, 360])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers tests
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_to_float_handles_currency(self):
        assert to_float("₹1,23,456.78") == 123456.78

    def test_to_float_handles_accounting_negatives(self):
        assert to_float("(1,200.50)") == -1200.50

    def test_to_float_handles_nan_and_blanks(self):
        assert to_float(None) == 0.0
        assert to_float("") == 0.0
        assert to_float("nan") == 0.0
        assert to_float(float("nan")) == 0.0

    def test_to_int_coerces(self):
        assert to_int("42") == 42
        assert to_int("42.9") == 42
        assert to_int("not-a-number") == 0


# ---------------------------------------------------------------------------
# Flipkart parser tests
# ---------------------------------------------------------------------------

class TestFlipkartParser:
    @pytest.fixture
    def parsed(self) -> dict:
        return parse_flipkart_settlement(_make_flipkart_workbook())

    def test_summary_extracted(self, parsed):
        s = parsed["summary"]
        assert s["payment_duration"] == "01-Apr to 15-Apr"
        assert s["total_sale_orders"] == 120
        assert s["total_returns"] == 8
        assert s["gross_sales_amount"] == 150000.0
        assert s["returns_reversal"] == -12000.0
        assert s["marketplace_fees"] == -15000.0
        assert s["tcs_amount"] == -1500.0
        assert s["net_bank_settlement"] == 112800.0
        assert s["total_realizable_amount"] == 116500.0

    def test_orders_extracted(self, parsed):
        orders = parsed["orders"]
        assert len(orders) == 3
        first = orders[0]
        assert first["order_id"] == "OD111"
        assert first["sale_amount"] == 1500.0
        assert first["seller_sku"] == "SKU-A"
        assert first["quantity"] == 1
        assert first["reverse_shipping_fee"] == 0.0

    def test_ads_total_spend(self, parsed):
        # 3000 + 540 + 2000 + 360 = 5900
        assert parsed["ads_total_spend"] == 5900.0
        assert len(parsed["ads"]) == 2

    def test_sku_aggregation(self, parsed):
        skus = {s["seller_sku"]: s for s in parsed["skus"]}
        assert "SKU-A" in skus
        assert "SKU-B" in skus

        sku_a = skus["SKU-A"]
        # Two orders for SKU-A: one normal (qty 1, sale 1500), one return (refund -1500)
        assert sku_a["total_orders"] == 2
        assert sku_a["return_orders"] == 1
        assert sku_a["return_rate"] == 50.0
        assert sku_a["total_revenue"] == 3000.0  # 1500 + 1500
        assert sku_a["total_refunds"] == -1500.0

        sku_b = skus["SKU-B"]
        assert sku_b["units_sold"] == 2
        assert sku_b["return_orders"] == 0

    def test_no_parsing_errors_on_clean_input(self, parsed):
        # Per-row errors should be empty for a clean workbook
        critical = [e for e in parsed["parsing_errors"] if "row" in e.lower()]
        assert critical == []

    def test_handles_invalid_bytes(self):
        result = parse_flipkart_settlement(b"not an excel file")
        assert result["summary"] == {}
        assert result["orders"] == []
        assert any("Failed to open" in e for e in result["parsing_errors"])


# ---------------------------------------------------------------------------
# Amazon parser tests
# ---------------------------------------------------------------------------

class TestAmazonParser:
    def test_amazon_csv_basic(self):
        csv = (
            "order id,sku,quantity purchased,product sales,selling fees,"
            "taxes,refund amount,total,transaction type,posted date time\n"
            "AMZ1,SKU-X,1,500,-50,90,0,540,Order,2026-04-01\n"
            "AMZ2,SKU-X,2,1000,-100,180,0,1080,Order,2026-04-02\n"
            "AMZ3,SKU-Y,1,800,-80,144,-800,-80,Refund,2026-04-03\n"
            "AMZ4,,0,0,0,0,-300,-300,Sponsored Ads,2026-04-04\n"
        )
        result = parse_amazon_settlement(csv.encode())

        assert result["platform"] == "amazon"
        # 3 order rows; the ads row is routed to ads bucket
        assert len(result["orders"]) == 3
        assert len(result["ads"]) == 1
        assert result["ads_total_spend"] == -300.0

        # Summary synthesised
        assert result["summary"]["total_sale_orders"] == 3
        assert result["summary"]["total_returns"] == 1
        assert result["summary"]["gross_sales_amount"] == 2300.0

        # SKU aggregation
        skus = {s["seller_sku"]: s for s in result["skus"]}
        assert skus["SKU-X"]["units_sold"] == 3
        assert skus["SKU-X"]["total_revenue"] == 1500.0
        assert skus["SKU-Y"]["return_orders"] == 1


# ---------------------------------------------------------------------------
# Direct unit tests for SKU aggregator
# ---------------------------------------------------------------------------

class TestAggregateSkus:
    def test_empty_input(self):
        assert aggregate_skus([]) == []

    def test_groups_and_computes_metrics(self):
        orders = [
            {"seller_sku": "S1", "sale_amount": 100, "marketplace_fee": -10,
             "refund_amount": 0, "reverse_shipping_fee": 0,
             "bank_settlement_value": 90, "quantity": 1},
            {"seller_sku": "S1", "sale_amount": 200, "marketplace_fee": -20,
             "refund_amount": -200, "reverse_shipping_fee": -30,
             "bank_settlement_value": -50, "quantity": 1},
        ]
        result = aggregate_skus(orders)
        assert len(result) == 1
        s1 = result[0]
        assert s1["total_orders"] == 2
        assert s1["return_orders"] == 1
        assert s1["return_rate"] == 50.0
        assert s1["units_sold"] == 2
        assert s1["avg_selling_price"] == 150.0
        assert s1["net_per_unit"] == 20.0  # (90 + -50) / 2


# ---------------------------------------------------------------------------
# Real-format Flipkart settlement (mirrors the live export layout)
# ---------------------------------------------------------------------------

def _make_real_format_flipkart_workbook() -> bytes:
    """Build a workbook in the *real* Flipkart layout used by sellers in production.

    Differences vs the legacy/sample layout:
      - Summary uses a structured table ('Line Item / Break-up L1 / Break-up L2 /
        Amount Settled (Rs.) / …'), with many break-up values being unresolved
        cross-sheet formulas (``#gid=…``) — the parser must skip those and rely
        on the group-level totals + Orders aggregates.
      - Orders columns include "(Rs.)" suffixes and trailing formula
        annotations ("\\n= SUM(V:AI)").
      - Ads sheet has a group-header row 0 + column row 1 with wallet-based
        accounting (``Wallet Topup`` / ``Wallet Refund`` / ``GST on Ads Fees``).
    """
    from openpyxl import Workbook

    wb = Workbook()

    # --- Summary --------------------------------------------------------------
    s = wb.active
    s.title = "Summary of report"
    s.append([None, "Settlement Report Summary"])
    s.append([None, "NOTE: Test fixture"])
    s.append([None, None])
    s.append([None, "Payment Duration ", "2026-04-01 - 2026-04-30"])
    s.append([None, "# Sale Orders (Not returned)", 3012])
    s.append([None, "# Returns", 909])
    s.append([None, None])
    s.append([
        None, "Line Item", "Break-up L1", "Break-up L2",
        "Amount Settled (Rs.)", "Break-up L1 (Rs.)", "Break-up L2 (Rs.)",
        "Remarks (Information about each line)",
    ])
    s.append([None, "Orders", None, None, 498693.65])
    # These rows have placeholder formulas — parser must skip them
    s.append([None, None, "Sales Amount", None, None, "#gid=1241438932&range=J2", None, "desc"])
    s.append([None, None, "Returns Reversal", None, None, "#gid=1241438932&range=R2", None, "desc"])
    s.append([None, None, "Marketplace Fees", None, None, "#gid=1241438932&range=N2", None, "desc"])
    s.append([None, None, "Taxes", None, -20232.48, None, None, "with breakup below"])
    s.append([None, None, None, "TCS", None, None, "#gid=1241438932&range=AK2", "desc"])
    s.append([None, "Net Bank Settlement (A)", None, None, 497583.09])
    s.append([None, "Input GST + TCS Credits (B)", None, None, 28384.39])
    s.append([None, "Income Tax Credits (C)", None, None, 583.03])
    s.append([None, "Total Realizable Amount", None, None, 526550.51])

    # --- Orders ---------------------------------------------------------------
    o = wb.create_sheet("Orders")
    o.append(["Payment Details"] * 6 + ["Order"] * 2 + ["Financials"] * 5 + ["Product"] * 4)
    o.append([
        "NEFT ID", "Neft Type", " Payment Date",
        "Bank Settlement Value (Rs.) \n= SUM(J:R)",
        "Input GST + TCS Credits (Rs.)\n[GST+TCS]",
        "Income Tax Credits (Rs.)\n[TDS]",
        "Order ID", "Order item ID",
        "Sale Amount (Rs.)", "Marketplace Fee (Rs.)\n= SUM (V:AI)",
        "Taxes (Rs.)", "Refund (Rs.)", "GST on MP Fees (Rs.)",
        "Seller SKU", "Quantity", "Product Sub Category", "Return Type",
    ])
    o.append(["NFT-1", "PREPAID", "2026-04-05", 270, 5, 5,
              "OD1", "OI1", 290, -8.03, -3.10, 0, -1.44, "BB-101-NBLU-01", 1, "baby", None])
    o.append(["NFT-2", "PREPAID", "2026-04-06", 0, 5, 5,
              "OD2", "OI2", 290, -8.03, -3.10, -290, -1.44, "BB-101-NBLU-01", 1, "baby", "Customer Return"])
    o.append(["NFT-3", "PREPAID", "2026-04-07", 1380, 25, 12,
              "OD3", "OI3", 1500, -45.0, -75.0, 0, -8.10, "SH-201-RED-M", 1, "shoes", None])

    # --- Ads (real format: group-header row 0, columns row 1) -----------------
    a = wb.create_sheet("Ads")
    a.append(["Payment Details", None, None, None, "Transaction Summary",
              None, None, None, None, None, None])
    a.append([
        "NEFT ID", "Payment Date", "Settlement Value (Rs.) = SUM(G:K)", None,
        "Type", "Campaign / Transaction ID",
        "Wallet Redeem (Rs.)", "Wallet Redeem Reversal (Rs.)",
        "Wallet Topup (Rs.)", "Wallet Refund (Rs.)", "GST on Ads Fees (Rs.)",
    ])
    a.append(["NFT-AD-1", "2026-04-10", 5000, None, "topup", "TXN-1",
              0, 0, 5000, 0, -900])
    a.append(["NFT-AD-2", "2026-04-15", 0, None, "redeem", "PCID-A1",
              -2000, 0, 0, 0, -360])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestRealFlipkartFormat:
    """Validate that the parser handles the live Flipkart settlement layout."""

    @pytest.fixture
    def parsed(self) -> dict:
        return parse_flipkart_settlement(_make_real_format_flipkart_workbook())

    def test_summary_top_line_totals(self, parsed):
        s = parsed["summary"]
        # Top-line totals come straight from the Line Item table
        assert s["payment_duration"] == "2026-04-01 - 2026-04-30"
        assert s["total_sale_orders"] == 3012
        assert s["total_returns"] == 909
        assert s["net_bank_settlement"] == 497583.09
        assert s["input_gst_tcs_credits"] == 28384.39
        assert s["income_tax_credits"] == 583.03
        assert s["total_realizable_amount"] == 526550.51

    def test_summary_breakup_backfilled_from_orders(self, parsed):
        # Summary break-up values were #gid placeholders → must be derived from Orders
        s = parsed["summary"]
        # Sum of sale_amount across the 3 order rows = 290 + 290 + 1500 = 2080
        assert s["gross_sales_amount"] == 2080.0
        # marketplace_fee column sums to -61.06
        assert s["marketplace_fees"] == round(-8.03 - 8.03 - 45.0, 2)
        # Returns reversal = -sum(refund < 0) = -(-290) = 290
        assert s["returns_reversal"] == 290.0

    def test_orders_columns_with_rs_suffix_resolve(self, parsed):
        orders = parsed["orders"]
        assert len(orders) == 3
        assert orders[0]["order_id"] == "OD1"
        assert orders[0]["seller_sku"] == "BB-101-NBLU-01"
        assert orders[0]["sale_amount"] == 290.0
        # "Marketplace Fee (Rs.)\n= SUM (V:AI)" must resolve correctly
        assert orders[0]["marketplace_fee"] == -8.03

    def test_ads_wallet_layout(self, parsed):
        # Two real ad rows (header row dropped)
        assert len(parsed["ads"]) == 2
        topup = next(a for a in parsed["ads"] if a["transaction_type"] == "topup")
        assert topup["amount"] == 5000.0  # Wallet Topup - Wallet Refund
        assert topup["gst_on_ads"] == 900.0  # absolute value
        # Total spend = (5000 - 0) + (0 - 0) + (900 + 360) = 6260
        assert parsed["ads_total_spend"] == 6260.0

    def test_no_parsing_errors(self, parsed):
        assert parsed["parsing_errors"] == []

    def test_sku_aggregation_with_returns(self, parsed):
        skus = {s["seller_sku"]: s for s in parsed["skus"]}
        # BB-101 has one normal + one returned order
        bb = skus["BB-101-NBLU-01"]
        assert bb["total_orders"] == 2
        assert bb["return_orders"] == 1
        assert bb["return_rate"] == 50.0
